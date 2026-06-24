# -*- coding: utf-8 -*-
import os
import json
import uuid
from datetime import datetime, date, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file, flash
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, inspect as sa_inspect
from werkzeug.security import generate_password_hash, check_password_hash
from apscheduler.schedulers.background import BackgroundScheduler
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

# ─────────────────────────────────────────────
# App Setup
# ─────────────────────────────────────────────
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
_default_db_dir = os.path.join(BASE_DIR, "database")
DB_DIR  = os.environ.get("DB_DIR", _default_db_dir)
DB_PATH = os.path.join(DB_DIR, "sk_egg_mart.db")
os.makedirs(DB_DIR, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "skeggmart-secret-2024-xK9p")
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# ─────────────────────────────────────────────
# Database Models
# ─────────────────────────────────────────────

class Customer(db.Model):
    __tablename__ = "customers"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    block_name = db.Column(db.String(5), nullable=False)
    room_number = db.Column(db.String(20), nullable=False)
    mobile_number = db.Column(db.String(15), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    orders = db.relationship("Order", backref="customer", lazy=True)

class Order(db.Model):
    __tablename__ = "orders"
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.String(30), unique=True, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=False)
    status = db.Column(db.String(30), default="Pending")
    subtotal_amount = db.Column(db.Float, default=0.0)
    delivery_charge = db.Column(db.Float, default=0.0)
    grand_total = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    items = db.relationship("OrderItem", backref="order", lazy=True, cascade="all, delete-orphan")

class OrderItem(db.Model):
    __tablename__ = "order_items"
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey("orders.id"), nullable=False)
    egg_type = db.Column(db.String(30), nullable=False)
    trays = db.Column(db.Integer, nullable=False)
    additional_eggs = db.Column(db.Integer, default=0)
    total_eggs = db.Column(db.Integer, nullable=False)
    egg_price = db.Column(db.Float, default=0.0)
    item_amount = db.Column(db.Float, default=0.0)

class DailySummary(db.Model):
    __tablename__ = "daily_summary"
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, unique=True, nullable=False)
    total_orders = db.Column(db.Integer, default=0)
    total_trays = db.Column(db.Integer, default=0)
    total_eggs = db.Column(db.Integer, default=0)
    white_eggs = db.Column(db.Integer, default=0)
    brown_eggs = db.Column(db.Integer, default=0)
    total_revenue = db.Column(db.Float, default=0.0)

class PricingSettings(db.Model):
    __tablename__ = "pricing_settings"
    id = db.Column(db.Integer, primary_key=True)
    white_egg_price = db.Column(db.Float, default=6.0)
    brown_egg_price = db.Column(db.Float, default=8.0)
    delivery_charge = db.Column(db.Float, default=20.0)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class AdminUser(db.Model):
    __tablename__ = "admin_users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def get_pricing():
    """Return current pricing settings (always returns a record)."""
    p = PricingSettings.query.first()
    if not p:
        p = PricingSettings(white_egg_price=6.0, brown_egg_price=8.0, delivery_charge=20.0)
        db.session.add(p)
        db.session.commit()
    return p

def generate_order_id():
    today = date.today().strftime("%Y%m%d")
    prefix = f"SKEGG-{today}-"
    last = Order.query.filter(Order.order_id.like(f"{prefix}%")).order_by(Order.id.desc()).first()
    if last:
        seq = int(last.order_id.split("-")[-1]) + 1
    else:
        seq = 1
    return f"{prefix}{str(seq).zfill(3)}"

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "admin_logged_in" not in session:
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated

def fmt_inr(amount):
    return f"₹{amount:,.2f}"

def order_to_dict(order):
    customer = order.customer
    items = []
    total_trays = 0
    total_eggs = 0
    for item in order.items:
        items.append({
            "egg_type": item.egg_type,
            "trays": item.trays,
            "additional_eggs": item.additional_eggs,
            "total_eggs": item.total_eggs,
            "egg_price": item.egg_price,
            "item_amount": item.item_amount,
        })
        total_trays += item.trays
        total_eggs += item.total_eggs
    return {
        "id": order.id,
        "order_id": order.order_id,
        "status": order.status,
        "created_at": order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        "date": order.created_at.strftime("%d %b %Y"),
        "time": order.created_at.strftime("%I:%M %p"),
        "customer": {
            "name": customer.name,
            "block_name": customer.block_name,
            "room_number": customer.room_number,
            "mobile_number": customer.mobile_number,
        },
        "items": items,           # for JSON API / JS
        "order_items": items,     # for Jinja2 templates
        "total_trays": total_trays,
        "total_eggs": total_eggs,
        "subtotal_amount": order.subtotal_amount or 0.0,
        "delivery_charge": order.delivery_charge or 0.0,
        "grand_total": order.grand_total or 0.0,
    }

def _add_col_if_missing(table, col, col_def):
    """Safely add a column to an existing SQLite table."""
    try:
        inspector = sa_inspect(db.engine)
        cols = [c["name"] for c in inspector.get_columns(table)]
        if col not in cols:
            with db.engine.connect() as conn:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {col} {col_def}"))
                conn.commit()
    except Exception:
        pass

# ─────────────────────────────────────────────
# Auto Daily Reset (APScheduler)
# ─────────────────────────────────────────────

def daily_reset_job():
    """Runs at midnight: archive yesterday's summary with revenue."""
    with app.app_context():
        today = date.today()
        yesterday = today - timedelta(days=1)
        existing = DailySummary.query.filter_by(date=yesterday).first()
        if existing:
            return
        start = datetime.combine(yesterday, datetime.min.time())
        end   = datetime.combine(today, datetime.min.time())
        orders = Order.query.filter(Order.created_at >= start, Order.created_at < end).all()
        total_orders = len(orders)
        total_trays = total_eggs = white_eggs = brown_eggs = 0
        total_revenue = 0.0
        for o in orders:
            total_revenue += o.grand_total or 0.0
            for item in o.items:
                total_trays += item.trays
                total_eggs  += item.total_eggs
                if "white" in item.egg_type.lower():
                    white_eggs += item.total_eggs
                else:
                    brown_eggs += item.total_eggs
        summary = DailySummary(
            date=yesterday,
            total_orders=total_orders,
            total_trays=total_trays,
            total_eggs=total_eggs,
            white_eggs=white_eggs,
            brown_eggs=brown_eggs,
            total_revenue=total_revenue,
        )
        db.session.add(summary)
        db.session.commit()

# ─────────────────────────────────────────────
# DB Initialization
# ─────────────────────────────────────────────

def init_db():
    db.create_all()
    # Migrate existing tables with new columns
    _add_col_if_missing("orders",      "subtotal_amount", "REAL DEFAULT 0")
    _add_col_if_missing("orders",      "delivery_charge", "REAL DEFAULT 0")
    _add_col_if_missing("orders",      "grand_total",     "REAL DEFAULT 0")
    _add_col_if_missing("order_items", "egg_price",       "REAL DEFAULT 0")
    _add_col_if_missing("order_items", "item_amount",     "REAL DEFAULT 0")
    _add_col_if_missing("daily_summary","total_revenue",  "REAL DEFAULT 0")
    # Default admin
    if not AdminUser.query.first():
        db.session.add(AdminUser(
            username="admin",
            password_hash=generate_password_hash("skegg@2024")
        ))
        db.session.commit()
        print("[OK] Default admin created: admin / skegg@2024")
    # Default pricing
    if not PricingSettings.query.first():
        db.session.add(PricingSettings(
            white_egg_price=6.0,
            brown_egg_price=8.0,
            delivery_charge=20.0,
        ))
        db.session.commit()
        print("[OK] Default pricing settings created")

# ─────────────────────────────────────────────
# Customer Routes
# ─────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/order", methods=["GET", "POST"])
def order():
    if request.method == "POST":
        data    = request.get_json()
        pricing = get_pricing()

        # Save customer
        customer = Customer(
            name=data["customer_name"],
            block_name=data["block_name"],
            room_number=data["room_number"],
            mobile_number=data["mobile_number"],
        )
        db.session.add(customer)
        db.session.flush()

        order_id  = generate_order_id()
        subtotal  = 0.0
        del_charge = pricing.delivery_charge

        new_order = Order(
            order_id=order_id,
            customer_id=customer.id,
            delivery_charge=del_charge,
        )
        db.session.add(new_order)
        db.session.flush()

        for item in data["items"]:
            trays      = int(item["trays"])
            additional = int(item["additional_eggs"])
            total      = trays * 30 + additional
            egg_type   = item["egg_type"]
            price_per  = (pricing.white_egg_price
                          if "white" in egg_type.lower()
                          else pricing.brown_egg_price)
            amount     = round(total * price_per, 2)
            subtotal  += amount
            oi = OrderItem(
                order_id=new_order.id,
                egg_type=egg_type,
                trays=trays,
                additional_eggs=additional,
                total_eggs=total,
                egg_price=price_per,
                item_amount=amount,
            )
            db.session.add(oi)

        new_order.subtotal_amount = round(subtotal, 2)
        new_order.grand_total     = round(subtotal + del_charge, 2)
        db.session.commit()
        return jsonify({"success": True, "order_id": order_id})
    return render_template("order.html")

@app.route("/order/success/<order_id>")
def order_success(order_id):
    order = Order.query.filter_by(order_id=order_id).first_or_404()
    return render_template("order_success.html", order=order_to_dict(order))

# ─────────────────────────────────────────────
# Favicon
# ─────────────────────────────────────────────
@app.route("/favicon.ico")
def favicon():
    return send_file(os.path.join(BASE_DIR, "static", "icons", "icon-192.png"),
                     mimetype="image/png")

# ─────────────────────────────────────────────
# Admin Routes
# ─────────────────────────────────────────────

@app.route("/admin", methods=["GET"])
def admin_index():
    if "admin_logged_in" in session:
        return redirect(url_for("admin_dashboard"))
    return redirect(url_for("admin_login"))

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        admin = AdminUser.query.filter_by(username=username).first()
        if admin and check_password_hash(admin.password_hash, password):
            session["admin_logged_in"] = True
            session["admin_username"] = username
            return redirect(url_for("admin_dashboard"))
        flash("Invalid credentials. Please try again.", "error")
    return render_template("admin/login.html")

@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))

@app.route("/admin/dashboard")
@login_required
def admin_dashboard():
    return render_template("admin/dashboard.html")

@app.route("/admin/orders")
@login_required
def admin_orders():
    return render_template("admin/orders.html")

@app.route("/admin/order/<order_id>")
@login_required
def admin_order_detail(order_id):
    order = Order.query.filter_by(order_id=order_id).first_or_404()
    o = order_to_dict(order)
    whatsapp_url = "https://wa.me/919445379128"
    return render_template("admin/order_detail.html", order=o, whatsapp_url=whatsapp_url)

@app.route("/admin/history")
@login_required
def admin_history():
    return render_template("admin/history.html")

@app.route("/admin/qr-code")
@login_required
def admin_qr_code():
    return render_template("admin/qr_code.html")

@app.route("/admin/pricing")
@login_required
def admin_pricing():
    pricing = get_pricing()
    return render_template("admin/pricing.html", pricing=pricing)

# ─────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────

@app.route("/api/pricing")
def api_get_pricing():
    """Public endpoint — customer order page fetches current prices."""
    p = get_pricing()
    return jsonify({
        "white_egg_price": p.white_egg_price,
        "brown_egg_price": p.brown_egg_price,
        "delivery_charge": p.delivery_charge,
    })

# ─────────────────────────────────────────────
# Admin API Routes
# ─────────────────────────────────────────────

@app.route("/api/pricing", methods=["POST"])
@login_required
def api_update_pricing():
    data = request.get_json()
    p = get_pricing()
    p.white_egg_price = float(data.get("white_egg_price", p.white_egg_price))
    p.brown_egg_price = float(data.get("brown_egg_price", p.brown_egg_price))
    p.delivery_charge = float(data.get("delivery_charge", p.delivery_charge))
    p.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"success": True,
                    "white_egg_price": p.white_egg_price,
                    "brown_egg_price": p.brown_egg_price,
                    "delivery_charge": p.delivery_charge})

@app.route("/api/dashboard-stats")
@login_required
def api_dashboard_stats():
    today  = date.today()
    start  = datetime.combine(today, datetime.min.time())
    orders_today = Order.query.filter(Order.created_at >= start).all()

    total_orders = len(orders_today)
    total_trays = total_eggs = white_eggs = brown_eggs = 0
    pending = delivered = 0
    today_revenue = 0.0

    for o in orders_today:
        today_revenue += o.grand_total or 0.0
        if o.status == "Pending":   pending   += 1
        if o.status == "Delivered": delivered += 1
        for item in o.items:
            total_trays += item.trays
            total_eggs  += item.total_eggs
            if "white" in item.egg_type.lower():
                white_eggs += item.total_eggs
            else:
                brown_eggs += item.total_eggs

    # Overall
    all_orders    = Order.query.all()
    overall_eggs  = sum(i.total_eggs for o in all_orders for i in o.items)
    overall_rev   = sum(o.grand_total or 0.0 for o in all_orders)

    # Yesterday revenue
    yesterday     = today - timedelta(days=1)
    ystart        = datetime.combine(yesterday, datetime.min.time())
    yend          = datetime.combine(today, datetime.min.time())
    yorders       = Order.query.filter(Order.created_at >= ystart, Order.created_at < yend).all()
    yesterday_rev = sum(o.grand_total or 0.0 for o in yorders)

    avg_order_val = round(overall_rev / len(all_orders), 2) if all_orders else 0.0

    return jsonify({
        "today_orders":       total_orders,
        "today_white_eggs":   white_eggs,
        "today_brown_eggs":   brown_eggs,
        "today_trays":        total_trays,
        "today_total_eggs":   total_eggs,
        "pending_orders":     pending,
        "delivered_orders":   delivered,
        "overall_total_eggs": overall_eggs,
        "today_revenue":      round(today_revenue, 2),
        "yesterday_revenue":  round(yesterday_rev, 2),
        "overall_revenue":    round(overall_rev, 2),
        "avg_order_value":    avg_order_val,
    })

@app.route("/api/orders")
@login_required
def api_orders():
    search      = request.args.get("search", "")
    status_filter = request.args.get("status", "")
    block_filter  = request.args.get("block", "")
    date_filter   = request.args.get("date", "")

    query = Order.query.join(Customer)

    if search:
        query = query.filter(
            db.or_(
                Customer.name.ilike(f"%{search}%"),
                Customer.mobile_number.ilike(f"%{search}%"),
                Order.order_id.ilike(f"%{search}%"),
            )
        )
    if status_filter:
        query = query.filter(Order.status == status_filter)
    if block_filter:
        query = query.filter(Customer.block_name == block_filter)
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
            start = datetime.combine(filter_date, datetime.min.time())
            end   = start + timedelta(days=1)
            query = query.filter(Order.created_at >= start, Order.created_at < end)
        except Exception:
            pass

    orders = query.order_by(Order.created_at.desc()).all()
    return jsonify([order_to_dict(o) for o in orders])

@app.route("/api/orders/latest-count")
@login_required
def api_latest_order_count():
    today = date.today()
    start = datetime.combine(today, datetime.min.time())
    count = Order.query.filter(Order.created_at >= start).count()
    return jsonify({"count": count})

@app.route("/api/orders/<order_id>/status", methods=["PUT"])
@login_required
def api_update_status(order_id):
    order = Order.query.filter_by(order_id=order_id).first_or_404()
    data  = request.get_json()
    order.status = data["status"]
    db.session.commit()
    return jsonify({"success": True, "status": order.status})

@app.route("/api/history")
@login_required
def api_history():
    days   = int(request.args.get("days", 7))
    result = []
    today  = date.today()
    for i in range(days):
        d     = today - timedelta(days=i)
        start = datetime.combine(d, datetime.min.time())
        end   = start + timedelta(days=1)
        orders = Order.query.filter(Order.created_at >= start, Order.created_at < end).all()
        total_orders = len(orders)
        total_trays  = sum(item.trays      for o in orders for item in o.items)
        total_eggs   = sum(item.total_eggs for o in orders for item in o.items)
        delivered    = sum(1 for o in orders if o.status == "Delivered")
        cancelled    = sum(1 for o in orders if o.status == "Cancelled")
        revenue      = sum(o.grand_total or 0.0 for o in orders)
        result.append({
            "date":         d.strftime("%d %b %Y"),
            "date_raw":     d.isoformat(),
            "total_orders": total_orders,
            "total_trays":  total_trays,
            "total_eggs":   total_eggs,
            "delivered":    delivered,
            "cancelled":    cancelled,
            "revenue":      round(revenue, 2),
        })
    return jsonify(result)

@app.route("/api/export-excel")
@login_required
def api_export_excel():
    orders = Order.query.join(Customer).order_by(Order.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SK EGG MART Orders"

    header_fill = PatternFill("solid", fgColor="F59E0B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    headers = [
        "Order ID", "Customer Name", "Block", "Room No", "Mobile",
        "Egg Type", "Trays", "Add.Eggs", "Total Eggs",
        "Egg Price (₹)", "Item Amount (₹)",
        "Subtotal (₹)", "Delivery (₹)", "Grand Total (₹)",
        "Status", "Date", "Time"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = thin

    col_widths = [18, 20, 8, 10, 14, 18, 8, 10, 12, 14, 16, 14, 12, 16, 14, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    alt_fill = PatternFill("solid", fgColor="FFF8E7")
    row_num = 2
    for o in orders:
        customer = o.customer
        for item in o.items:
            fill = alt_fill if row_num % 2 == 0 else None
            cells = [
                o.order_id, customer.name, customer.block_name, customer.room_number, customer.mobile_number,
                item.egg_type, item.trays, item.additional_eggs, item.total_eggs,
                item.egg_price, item.item_amount,
                o.subtotal_amount, o.delivery_charge, o.grand_total,
                o.status,
                o.created_at.strftime("%d %b %Y"),
                o.created_at.strftime("%I:%M %p"),
            ]
            ws.append(cells)
            for cell in ws[row_num]:
                cell.border = thin
                cell.alignment = center
                if fill:
                    cell.fill = fill
            row_num += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"SK_EGG_MART_Orders_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)

# ─────────────────────────────────────────────
# Scheduler + DB Init
# ─────────────────────────────────────────────

scheduler = BackgroundScheduler()
scheduler.add_job(daily_reset_job, "cron", hour=0, minute=0)
scheduler.start()

with app.app_context():
    init_db()

# ─────────────────────────────────────────────
# Entry
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("SK EGG MART Server Starting...")
    print("Customer Portal: http://localhost:5000")
    print("Admin Portal:    http://localhost:5000/admin")
    print("Admin Login:     admin / skegg@2024")
    port  = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV") != "production"
    app.run(debug=debug, host="0.0.0.0", port=port)
